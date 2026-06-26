from __future__ import annotations

import argparse
import csv
import re
import unicodedata
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


SHEET_CANDIDATES = ("Pct Watcher", "Registros")
BAD_DISCHARGE_WINDOW_DAYS = 10
FINAL_EVENT_TERMS = ("reintern", "hospitaliza", "internacao hospitalar", "internação hospitalar", "obito", "óbito")
ACUTE_TERMS = ("descompensacao aguda", "descompensação aguda")
FAVORABLE_OUTCOME_TERMS = (
    "melhora clinica",
    "melhora clínica",
    "condicao basal",
    "condição basal",
    "estabilizacao",
    "estabilização",
    "estabilizado",
    "melhora",
)
MONITORING_TERMS = ("em monitoramento", "monitorando")
BAD_DISCHARGE_TERMS = ("hospitalizacao", "hospitalização", "reinternacao", "reinternação", "internacao", "internação", "obito", "óbito")
TRANSFER_TERMS = ("transferencia para outro estabelecimento", "transferência para outro estabelecimento")
GOOD_DISCHARGE_TERMS = ("alta melhorada", "transf. entre unidades", "transferencia entre unidades", "migra", "alta a pedido", "evasao", "evasão")

ALIASES = {
    "record_date": ("data", "ultima_coleta", "data_hora"),
    "patient_name": ("paciente", "nome", "nome_do_paciente", "nome_paciente"),
    "unit": ("unidade", "unidade_assistencial", "setor"),
    "bed": ("leito",),
    "care_line": ("linha_de_cuidado",),
    "age": ("idade",),
    "risk_label": ("risco_ultimo", "risco"),
    "news2_last": ("news2_ultimo", "news2", "score_news2"),
    "news2_average_7d": ("news2_media_7_dias", "news2_m7d", "news2_media7dias"),
    "news2_delta_7d": ("delta_score_m7d", "delta_score_atual_e_medio", "delta_m7d"),
    "respiratory_rate": ("fr_irpm",),
    "oxygen_saturation": ("so_spo2",),
    "oxygen_support": ("sup_o2",),
    "systolic_bp": ("pas_mmhg",),
    "heart_rate": ("fc_bpm",),
    "consciousness": ("nc",),
    "temperature": ("temp_c",),
    "completeness": ("completude_de_ssvv", "ssvv_completeness"),
    "aura_alerted": ("alertado_aura",),
    "intervention_unit": ("intervencao_unidade", "intervencao_da_unidade"),
    "intervention_result": ("resultado_da_intervencao",),
    "clinical_outcome": ("desfecho_clinico",),
    "monitoring_status": ("status",),
    "committee_discussion": ("discussao_comite_aura",),
    "readmission_avoided": ("reinternacao_evitada",),
    "discharge_date": ("data_alta",),
    "clinical_alteration": ("alteracao_clinica",),
}

OUTPUT_COLUMNS = [
    "record_date",
    "record_datetime",
    "discharge_date",
    "patient_id",
    "patient_name",
    "unit",
    "care_line",
    "age",
    "news2_last",
    "news2_average_7d",
    "news2_delta_7d",
    "respiratory_rate",
    "oxygen_saturation",
    "oxygen_support_flag",
    "systolic_bp",
    "heart_rate",
    "consciousness_flag",
    "temperature",
    "completeness",
    "aura_alerted_flag",
    "acute_decompensation_flag",
    "deterioration_reversed_flag",
    "no_return_flag",
    "bad_discharge_within_10d",
    "bad_discharge_condition",
    "bad_discharge_date",
    "triage_score",
    "triage_band",
    "target_label",
    "target_readmission_event",
    "target_effective_intervention",
]


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Build the supervised AURA committee training dataset from Patient Watcher workbook."
    )
    parser.add_argument("workbook", nargs="?", type=Path, help="Path to .xlsx/.xls file.")
    parser.add_argument("--out", type=Path, default=Path("data/training_dataset.csv"))
    args = parser.parse_args()

    workbook_path = args.workbook or find_default_workbook()
    if workbook_path is None:
        raise SystemExit("No workbook found. Pass the .xlsx path as the first argument.")

    dataset = build_dataset(workbook_path)
    args.out.parent.mkdir(parents=True, exist_ok=True)

    with args.out.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        writer.writerows(dataset)

    labels = {}
    for row in dataset:
        labels[row["target_label"]] = labels.get(row["target_label"], 0) + 1

    print(f"Workbook: {workbook_path}")
    print(f"Rows: {len(dataset)}")
    print(f"Output: {args.out}")
    print(f"Labels: {labels}")


def find_default_workbook() -> Path | None:
    roots = [Path.cwd(), Path.cwd().parent]
    for root in roots:
        matches = sorted(root.glob("*.xlsx")) + sorted(root.glob("*.xls"))
        if matches:
            return matches[0]
    return None


def build_dataset(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_name = next((name for name in SHEET_CANDIDATES if name in workbook.sheetnames), workbook.sheetnames[0])
    worksheet = workbook[sheet_name]
    bad_events = build_bad_event_index(workbook)

    rows = worksheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return []

    keys = [normalize_key(str(value)) if value is not None else "" for value in headers]
    dataset: list[dict[str, Any]] = []

    for index, values in enumerate(rows, start=1):
        row = {key: value for key, value in zip(keys, values) if key}
        if not row or not pick(row, ALIASES["patient_name"]):
            continue
        dataset.append(build_training_row(row, index, bad_events))

    return dataset


def build_training_row(row: dict[str, Any], index: int, bad_events: dict[str, list[dict[str, Any]]]) -> dict[str, Any]:
    record_dt = parse_datetime(pick(row, ALIASES["record_date"]))
    patient_name = text(row, ALIASES["patient_name"]) or f"Paciente {index}"
    unit = text(row, ALIASES["unit"]) or "Sem unidade"
    care_line = text(row, ALIASES["care_line"]) or ""
    age = number(row, ALIASES["age"])
    risk_label = text(row, ALIASES["risk_label"])
    news2_last = number(row, ALIASES["news2_last"])
    news2_average_7d = number(row, ALIASES["news2_average_7d"])
    news2_delta_7d = number(row, ALIASES["news2_delta_7d"])
    respiratory_rate = number(row, ALIASES["respiratory_rate"])
    oxygen_saturation = number(row, ALIASES["oxygen_saturation"])
    oxygen_support = text(row, ALIASES["oxygen_support"])
    systolic_bp = number(row, ALIASES["systolic_bp"])
    heart_rate = number(row, ALIASES["heart_rate"])
    consciousness = text(row, ALIASES["consciousness"])
    temperature = number(row, ALIASES["temperature"])
    completeness = number(row, ALIASES["completeness"])
    aura_alerted = text(row, ALIASES["aura_alerted"])
    intervention_unit = text(row, ALIASES["intervention_unit"])
    intervention_result = text(row, ALIASES["intervention_result"])
    clinical_outcome = text(row, ALIASES["clinical_outcome"])
    monitoring_status = text(row, ALIASES["monitoring_status"])
    committee_discussion = text(row, ALIASES["committee_discussion"])
    readmission_avoided = text(row, ALIASES["readmission_avoided"])
    discharge_dt = parse_datetime(pick(row, ALIASES["discharge_date"]))
    clinical_alteration = text(row, ALIASES["clinical_alteration"])
    acute_decompensation = is_acute_decompensation(clinical_alteration)
    deterioration_reversed = is_deterioration_reversal(
        clinical_alteration=clinical_alteration,
        clinical_outcome=clinical_outcome,
        intervention_result=intervention_result,
    )
    no_return = contains_any(" | ".join([intervention_unit or "", intervention_result or "", clinical_outcome or ""]), ("sem retorno",))
    bad_event = find_bad_event_for_row(patient_name, record_dt, bad_events)

    score = score_case(
        news2_last=news2_last,
        news2_average_7d=news2_average_7d,
        news2_delta_7d=news2_delta_7d,
        respiratory_rate=respiratory_rate,
        oxygen_saturation=oxygen_saturation,
        oxygen_support=oxygen_support,
        systolic_bp=systolic_bp,
        heart_rate=heart_rate,
        consciousness=consciousness,
        temperature=temperature,
        completeness=completeness,
        aura_alerted=aura_alerted,
        risk_label=risk_label,
        intervention_unit=intervention_unit,
        monitoring_status=monitoring_status,
        readmission_avoided=readmission_avoided,
        discharge_date=discharge_dt.isoformat(sep=" ") if discharge_dt else None,
        acute_decompensation=acute_decompensation,
        no_return=no_return,
    )
    label = derive_training_label(
        intervention_result=intervention_result,
        clinical_outcome=clinical_outcome,
        monitoring_status=monitoring_status,
        committee_discussion=committee_discussion,
        readmission_avoided=readmission_avoided,
        intervention_unit=intervention_unit,
        discharge_date=discharge_dt.isoformat(sep=" ") if discharge_dt else None,
        clinical_alteration=clinical_alteration,
        bad_discharge=bad_event is not None,
        deterioration_reversed=deterioration_reversed,
    )

    return {
        "record_date": record_dt.date().isoformat() if record_dt else "",
        "record_datetime": record_dt.isoformat(sep=" ") if record_dt else "",
        "discharge_date": discharge_dt.date().isoformat() if discharge_dt else "",
        "patient_id": f"{patient_name}-{index}",
        "patient_name": patient_name,
        "unit": unit,
        "care_line": care_line,
        "age": age,
        "news2_last": news2_last,
        "news2_average_7d": news2_average_7d,
        "news2_delta_7d": news2_delta_7d,
        "respiratory_rate": respiratory_rate,
        "oxygen_saturation": oxygen_saturation,
        "oxygen_support_flag": 1 if oxygen_support and not contains_any(oxygen_support, ("nao", "não", "ar ambiente")) else 0,
        "systolic_bp": systolic_bp,
        "heart_rate": heart_rate,
        "consciousness_flag": 1 if consciousness and not contains_any(consciousness, ("alerta", "consciente", "normal")) else 0,
        "temperature": temperature,
        "completeness": completeness,
        "aura_alerted_flag": 1 if is_yes(aura_alerted) else 0,
        "acute_decompensation_flag": 1 if acute_decompensation else 0,
        "deterioration_reversed_flag": 1 if deterioration_reversed else 0,
        "no_return_flag": 1 if no_return else 0,
        "bad_discharge_within_10d": 1 if bad_event else 0,
        "bad_discharge_condition": bad_event["condition"] if bad_event else "",
        "bad_discharge_date": bad_event["date"].date().isoformat() if bad_event else "",
        "triage_score": round(score, 1),
        "triage_band": band_from_score(score),
        "target_label": label,
        "target_readmission_event": 1 if bad_event else 0,
        "target_effective_intervention": 1 if label in ("reinternacao_evitada", "reversao_piora") else 0,
    }


def score_case(**kwargs: Any) -> float:
    score = 0.0
    news2_last = kwargs["news2_last"]
    news2_average_7d = kwargs["news2_average_7d"]

    if news2_last is not None:
        score += news2_last * 1.15
    if news2_last is not None and news2_average_7d is not None:
        delta = news2_last - news2_average_7d
        score += delta * 1.8 if delta > 0 else -min(abs(delta), 2) * 0.6
    if kwargs["news2_delta_7d"] is not None:
        score += kwargs["news2_delta_7d"] * 1.1
    if is_yes(kwargs["aura_alerted"]):
        score += 1.5
    if contains_any(kwargs["risk_label"], ("alto", "critico", "crítico")):
        score += 1.4
    elif contains_any(kwargs["risk_label"], ("medio", "médio")):
        score += 0.6
    if kwargs["respiratory_rate"] is not None and kwargs["respiratory_rate"] >= 24:
        score += 1.5
    if kwargs["oxygen_saturation"] is not None and kwargs["oxygen_saturation"] < 92:
        score += 1.8
    if kwargs["systolic_bp"] is not None and (kwargs["systolic_bp"] <= 90 or kwargs["systolic_bp"] >= 180):
        score += 1.2
    if kwargs["heart_rate"] is not None and (kwargs["heart_rate"] >= 120 or kwargs["heart_rate"] <= 45):
        score += 1.2
    if kwargs["temperature"] is not None and (kwargs["temperature"] >= 38 or kwargs["temperature"] <= 35.5):
        score += 0.9
    if kwargs["consciousness"] and not contains_any(kwargs["consciousness"], ("alerta", "consciente", "normal")):
        score += 1.1
    if kwargs["completeness"] is not None and kwargs["completeness"] < 0.7:
        score -= 0.8
    if kwargs["acute_decompensation"]:
        score += 1.6
    if is_yes(kwargs["readmission_avoided"]):
        score -= 1.1
    if kwargs["no_return"] or contains_any(kwargs["intervention_unit"], ("sem retorno",)):
        score += 1.2
    return score


def derive_training_label(**kwargs: Any) -> str:
    if kwargs["bad_discharge"]:
        return "reinternacao_inevitavel"
    if kwargs["deterioration_reversed"]:
        return "reversao_piora"
    combined = normalize_text(" | ".join(str(value or "") for value in kwargs.values()))
    if is_yes(kwargs["readmission_avoided"]) or "evit" in combined:
        return "reinternacao_evitada"
    if any(term in combined for term in ("revers", "melhora", "estavel", "estável")):
        return "reversao_piora"
    if "sem retorno" in combined:
        return "sem_retorno"
    if contains_any(combined, FINAL_EVENT_TERMS):
        return "reinternacao_inevitavel"
    return "monitoramento_ativo"


def build_bad_event_index(workbook: Any) -> dict[str, list[dict[str, Any]]]:
    events: list[dict[str, Any]] = []
    if "Pcte Altana" in workbook.sheetnames:
        events.extend(parse_altana_events(workbook["Pcte Altana"]))
    if "Reinternações" in workbook.sheetnames:
        events.extend(parse_generic_reinternation_events(workbook["Reinternações"]))
    if "Reinternacoes" in workbook.sheetnames:
        events.extend(parse_generic_reinternation_events(workbook["Reinternacoes"]))

    indexed: dict[str, list[dict[str, Any]]] = {}
    seen: set[tuple[str, str, str]] = set()
    for event in events:
        key = normalize_name(event["patient_name"])
        if not key or not event["date"]:
            continue
        signature = (key, event["date"].date().isoformat(), event["condition"])
        if signature in seen:
            continue
        seen.add(signature)
        indexed.setdefault(key, []).append(event)
    for values in indexed.values():
        values.sort(key=lambda item: item["date"])
    return indexed


def parse_altana_events(worksheet: Any) -> list[dict[str, Any]]:
    rows = worksheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return []
    keys = [normalize_key(str(value)) if value is not None else "" for value in headers]
    events: list[dict[str, Any]] = []

    for values in rows:
        row = {key: value for key, value in zip(keys, values) if key}
        patient_name = text(row, ("nome", "paciente"))
        event_date = parse_datetime(pick(row, ("dt_alta_hosp", "dt_alta_medica", "data_alta")))
        reason = text(row, ("motivo", "condicao_alta", "condição_alta"))
        condition = classify_bad_discharge(reason)
        if patient_name and event_date and condition:
            events.append({"patient_name": patient_name, "date": event_date, "condition": condition, "source": "Pcte Altana"})
    return events


def parse_generic_reinternation_events(worksheet: Any) -> list[dict[str, Any]]:
    rows = worksheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return []
    keys = [normalize_key(str(value)) if value is not None else "" for value in headers]
    events: list[dict[str, Any]] = []

    for values in rows:
        row = {key: value for key, value in zip(keys, values) if key}
        patient_name = text(row, ("nome", "paciente", "patient_name"))
        event_date = parse_datetime(pick(row, ("data_alta", "discharge_date", "dt_alta_hosp", "dt_alta_medica")))
        condition_text = text(row, ("condicao_alta", "condição_alta", "motivo", "status"))
        condition = classify_bad_discharge(condition_text)
        if patient_name and event_date and condition:
            events.append({"patient_name": patient_name, "date": event_date, "condition": condition, "source": worksheet.title})
    return events


def classify_bad_discharge(value: str | None) -> str | None:
    normalized = normalize_text(value)
    if not normalized:
        return None
    if "obito" in normalized:
        return "Óbito"
    if any(term in normalized for term in TRANSFER_TERMS):
        return "Hospitalização"
    if any(term in normalized for term in BAD_DISCHARGE_TERMS) and not any(term in normalized for term in GOOD_DISCHARGE_TERMS):
        return "Hospitalização"
    return None


def find_bad_event_for_row(patient_name: str, record_dt: datetime | None, bad_events: dict[str, list[dict[str, Any]]]) -> dict[str, Any] | None:
    if not record_dt:
        return None
    events = bad_events.get(normalize_name(patient_name), [])
    for event in events:
        days = (event["date"].date() - record_dt.date()).days
        if 0 <= days <= BAD_DISCHARGE_WINDOW_DAYS:
            return event
    return None


def is_acute_decompensation(value: str | None) -> bool:
    return contains_any(value, ACUTE_TERMS)


def is_monitoring_outcome(value: str | None) -> bool:
    normalized = normalize_text(value)
    return bool(normalized) and "fim do monitoramento" not in normalized and any(term in normalized for term in MONITORING_TERMS)


def is_favorable_outcome(*values: str | None) -> bool:
    combined = normalize_text(" | ".join(value or "" for value in values))
    return any(term in combined for term in FAVORABLE_OUTCOME_TERMS)


def is_deterioration_reversal(*, clinical_alteration: str | None, clinical_outcome: str | None, intervention_result: str | None) -> bool:
    return (
        is_acute_decompensation(clinical_alteration)
        and is_favorable_outcome(clinical_outcome, intervention_result)
        and not is_monitoring_outcome(clinical_outcome)
    )


def band_from_score(score: float) -> str:
    if score >= 10:
        return "critico"
    if score >= 7:
        return "alto"
    if score >= 4:
        return "medio"
    return "baixo"


def normalize_key(value: str) -> str:
    normalized = unicodedata.normalize("NFD", value.strip().lower())
    normalized = "".join(char for char in normalized if unicodedata.category(char) != "Mn")
    normalized = re.sub(r"[\s\-/()]+", "_", normalized)
    normalized = re.sub(r"[^a-z0-9_]", "", normalized)
    normalized = re.sub(r"_+", "_", normalized).strip("_")
    return normalized


def pick(row: dict[str, Any], aliases: tuple[str, ...]) -> Any:
    for alias in aliases:
        value = row.get(alias)
        if value is not None and str(value).strip() != "":
            return value
    return None


def text(row: dict[str, Any], aliases: tuple[str, ...]) -> str | None:
    value = pick(row, aliases)
    if value is None:
        return None
    stripped = str(value).strip()
    return stripped or None


def number(row: dict[str, Any], aliases: tuple[str, ...]) -> float | None:
    value = pick(row, aliases)
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"-?\d+(?:[,.]\d+)?", str(value))
    if not match:
        return None
    return float(match.group(0).replace(",", "."))


def is_yes(value: str | None) -> bool:
    if not value:
        return False
    return normalize_text(value) in {"sim", "s", "yes", "y", "true", "1", "alerta", "alto", "critico", "crítico"}


def contains_any(value: str | None, terms: tuple[str, ...]) -> bool:
    if not value:
        return False
    lowered = normalize_text(value)
    return any(normalize_text(term) in lowered for term in terms)


def normalize_text(value: str | None) -> str:
    if not value:
        return ""
    normalized = unicodedata.normalize("NFD", str(value).strip().lower())
    return "".join(char for char in normalized if unicodedata.category(char) != "Mn")


def normalize_name(value: str | None) -> str:
    return re.sub(r"\s+", " ", normalize_text(value)).strip()


def parse_datetime(value: Any) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time.min)
    if isinstance(value, (int, float)) and value > 30000:
        try:
            return from_excel(value)
        except (TypeError, ValueError):
            return None
    text_value = str(value).strip()
    if not text_value:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y"):
        try:
            return datetime.strptime(text_value, fmt)
        except ValueError:
            continue
    return None


if __name__ == "__main__":
    main()
